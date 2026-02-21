import pandas as pd
import matplotlib.pyplot as plt
import os
import sys
from pathlib import Path
from datetime import datetime

# ========== 配置 ==========
INPUT_CSV = "./鹰眼记录.csv"
OUTPUT_EXCEL = "./缺陷日报.xlsx"
OUTPUT_CHART = "./缺陷趋势图.png"
# =========================

def setup_matplotlib():
    """配置 matplotlib 以支持中文显示"""
    system = sys.platform
    if system == 'win32':
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei']
    elif system == 'darwin':
        plt.rcParams['font.sans-serif'] = ['STHeiti', 'SimHei']
    else:
        plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False

def main():
    setup_matplotlib()
    
    input_path = Path(INPUT_CSV)
    if not input_path.exists():
        print(f"❌ 找不到文件：{input_path.resolve()}")
        return

    try:
        # 尝试多种编码读取 CSV
        encodings = ['utf-8-sig', 'utf-8', 'gbk']
        df = None
        for enc in encodings:
            try:
                df = pd.read_csv(input_path, encoding=enc)
                print(f"✅ 成功读取 CSV (编码：{enc})")
                break
            except UnicodeDecodeError:
                continue
        
        if df is None:
            print("❌ 无法解析 CSV 文件编码")
            return

        if df.empty:
            print("⚠️ CSV 文件为空")
            return

        # 检查必要列
        required_cols = ['时间', '缺陷数量']
        if not all(col in df.columns for col in required_cols):
            print(f"❌ CSV 缺少必要列，需要：{required_cols}")
            print(f"   当前列：{list(df.columns)}")
            return

        # ========== 数据清洗 ==========
        df['时间'] = pd.to_datetime(df['时间'], errors='coerce')
        df['缺陷数量'] = pd.to_numeric(df['缺陷数量'], errors='coerce').fillna(0)
        df = df.dropna(subset=['时间'])

        # 按日期聚合
        df['日期'] = df['时间'].dt.date
        daily = df.groupby('日期')['缺陷数量'].sum().reset_index()
        daily.columns = ['日期', '缺陷总数']
        
        # ⭐ 关键：将日期列转换为 datetime 类型（不是字符串！）
        daily['日期'] = pd.to_datetime(daily['日期'])
        
        # 验证数据类型
        print(f"📊 日期列类型：{daily['日期'].dtype}")
        print(f"📊 数据行数：{len(daily)}")
        print(daily.head())

        # ========== 写入 Excel (关键修复) ==========
        from openpyxl import Workbook
        from openpyxl.styles import numbers
        
        # 方法 1：使用 openpyxl 直接写入（最可靠）
        wb = Workbook()
        ws = wb.active
        ws.title = '缺陷日报'
        
        # 写入表头
        ws['A1'] = '日期'
        ws['B1'] = '缺陷总数'
        
        # 写入数据
        for idx, row in daily.iterrows():
            row_num = idx + 2
            # ⭐ 关键：直接写入 datetime 对象，不是字符串
            ws.cell(row=row_num, column=1, value=row['日期'])
            ws.cell(row=row_num, column=2, value=row['缺陷总数'])
            # ⭐ 关键：设置单元格格式为日期
            ws.cell(row=row_num, column=1).number_format = 'yyyy-mm-dd'
        
        # 设置表头格式
        from openpyxl.styles import Font, Alignment
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        
        wb.save(OUTPUT_EXCEL)
        print(f"✅ 日报已生成：{os.path.abspath(OUTPUT_EXCEL)}")

        # ========== 生成趋势图 ==========
        if len(daily) > 1:
            plt.figure(figsize=(10, 6))
            plt.plot(daily['日期'], daily['缺陷总数'], marker='o', linestyle='-', color='b')
            plt.title('每日缺陷趋势图')
            plt.xlabel('日期')
            plt.ylabel('缺陷总数 (个)')
            plt.grid(True, linestyle='--', alpha=0.6)
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig(OUTPUT_CHART, dpi=300)
            plt.close()
            print(f"✅ 趋势图已生成：{os.path.abspath(OUTPUT_CHART)}")
        else:
            print("ℹ️ 数据不足两天，跳过趋势图生成")

        # ========== 验证 Excel 日期格式 ==========
        print("\n🔍 验证 Excel 日期格式...")
        from openpyxl import load_workbook
        wb_verify = load_workbook(OUTPUT_EXCEL)
        ws_verify = wb_verify.active
        if ws_verify.max_row >= 2:
            cell_value = ws_verify['A2'].value
            cell_format = ws_verify['A2'].number_format
            print(f"   A2 单元格值：{cell_value} (类型：{type(cell_value).__name__})")
            print(f"   A2 单元格格式：{cell_format}")
            if isinstance(cell_value, datetime):
                print("   ✅ 日期格式正确！")
            else:
                print("   ⚠️ 日期可能仍为文本格式")
        wb_verify.close()

    except Exception as e:
        print("="*50)
        print("❌ 处理过程中发生错误:")
        print(str(e))
        import traceback
        traceback.print_exc()
        print("="*50)

if __name__ == "__main__":
    main()
